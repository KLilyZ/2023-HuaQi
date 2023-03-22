import openpyxl
import requests
import os
import shutil

#这里加一段删除前面数据的或者另存到另一个包里的代码

def RemoveDir(filepath):
    '''
    文件夹不存在就创建，存在就清空
    '''
    if not os.path.exists(filepath):
        os.mkdir(filepath)
    else:
        shutil.rmtree(filepath)
        os.mkdir(filepath)

root=r'..\DATA SUPPORT\词频统计\年报pdf版本'#存放pdf年报的位置
root2=r'..\DATA SUPPORT\词频统计\年报txt版本'#存放txt年报的位置
start=2#merged_data.xlsx文件中的第几行开始
end=3000#merged_data.xlsx文件中的第几行结束
RemoveDir(root)
RemoveDir(root2)

wb=openpyxl.load_workbook('merged_data.xlsx')
ws=wb.active
for i in range(start,end+1):
    try:
        dir_name=ws.cell(i,1).value+'-'+ws.cell(i,2).value
        pdf_name=ws.cell(i,3).value+'年度报告'

        dir_path=root+'\\'+dir_name
        pdf_path=dir_path+'\\'+pdf_name+'.pdf'
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
        pdf_url=ws.cell(i,4).value
        response=requests.get(pdf_url)
        data=response.content
        with open(pdf_path,mode='wb') as f:
            f.write(data)
        print(f'正在下载：{pdf_path}')
    except:
        ws.cell(i,5).value='有错误，请检查'
wb.save('merged_data.xlsx')
