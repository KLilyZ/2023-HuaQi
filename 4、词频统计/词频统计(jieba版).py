import jieba
from nltk.corpus import stopwords
import openpyxl

wb=openpyxl.load_workbook('jieba词频统计.xlsx')
ws=wb.active

punc = stopwords.words('biaodian')  # 要去除的中文标点符号
baidu_stopwords = stopwords.words('baidu_stopwords') # 导入停用词表
spic_list=['□','√',':','@','%','≥','≤','+','-','•','\n','\r',' ']
stop_word=baidu_stopwords+punc+spic_list

#判断是否为英文
def is_eng(txt):
    if (u'\u0041' <= txt <= u'\u005a') or (u'\u0061' <= txt <= u'\u007a'):
        return True
    else:
        return False

#判断是否为数字
def is_number(s):
    try:  # 如果能运行float(s)语句，返回True（字符串s是浮点数）
        float(s)
        return True
    except ValueError:  # ValueError为Python的一种标准异常，表示"传入无效的参数"
        pass  #如果引发了ValueError这种异常，不做任何事情（pass：不做任何事情，一般用做占位语句）
    try:
        import unicodedata  #处理ASCii码的包
        unicodedata.numeric(s)  #把一个表示数字的字符串转换为浮点数返回的函数
        return True
    except (TypeError,ValueError):
        pass
    return False

# 读入文件
with open('2016年年度报告.txt', encoding="utf-8") as fp:
    text = fp.read()
ls = jieba.lcut(text)#精确模式  #分词
word_list = []
for word in ls:
    if (word not in stop_word) and (not is_eng(word)) and (not is_number(word)) and (not (len(word)<=1 or len(word)>=15))and ('.' not in word ) and ('%' not in word):
        word_list.append(word.strip())
ws.cell(2,1).value=len(word_list)
for num,word in enumerate(word_list):
    print(word)
    ws.cell(num+3,1).value=word
wb.save('jieba词频统计.xlsx')
